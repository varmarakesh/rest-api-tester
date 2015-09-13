__author__ = 'rakesh.varma'
import openpyxl
from openpyxl import Workbook
from testDataLib import *
import csv
import os

class excelDataFileLib:
    testInputDataFile = None
    sheet = None
    testDataInputs = []

    def __init__(self, testInputDataFile):
        self.testInputDataFile = testInputDataFile
        wb = openpyxl.load_workbook(self.testInputDataFile)
        self.sheet = wb.get_active_sheet()


    def loadTestData(self):
        self.testDataInputs = []
        for x in range(2, self.sheet.get_highest_row()+1):
            testData = TestDataInput(url=self.sheet.cell(row = x, column = 1).value,
                                     input=self.sheet.cell(row = x, column = 2).value,
                                     expectedOutput=self.sheet.cell(row = x, column = 3).value,
                                     expectedHttpStatus=self.sheet.cell(row = x, column = 4).value,
                                     expectedError=self.sheet.cell(row = x, column = 5).value,
                                     testSummary=self.sheet.cell(row = x, column = 6).value,
                                     testDescription=self.sheet.cell(row = x, column = 7).value
                                     )
            self.testDataInputs.append(testData)

    def _writeToCell(self, worksheet, rowNum, columnNum, val):
        worksheet.cell(row = rowNum, column = columnNum).value = val

    def getTestData(self):
        return self.testDataInputs

    def getResultsFilePath(self):
        dir, file = os.path.split(self.testInputDataFile)
        prefix = file[0:file.rindex('.')]
        extension = file[file.rindex('.'): file.__len__()]
        results_file = prefix + '_results' + extension
        results_dir = os.path.split(os.path.dirname(self.testInputDataFile))[0] + '/results/'

        return results_dir + results_file

    def writeTestDataResults(self, testDataResults):
        wb = Workbook()
        ws = wb.active
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 1, val = 'URL')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 2, val = 'Input')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 3, val = 'Expected_Output')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 4, val = 'Expected_HTTP_Status')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 5, val = 'Expected_Error')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 6, val = 'Actual_Output')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 7, val = 'Actual_HTTP_Status')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 8, val = 'Actual_Error')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 9, val = 'Result')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 10, val = 'TestSummary')
        self._writeToCell(worksheet = ws, rowNum = 1, columnNum = 11, val = 'TestDescription')


        for index in range(1, len(testDataResults)):
            ws.cell(row = index+1, column = 1).value = testDataResults[index].url
            ws.cell(row = index+1, column = 2).value = testDataResults[index].input
            ws.cell(row = index+1, column = 3).value = testDataResults[index].expectedOutput
            ws.cell(row = index+1, column = 4).value = testDataResults[index].expectedHttpStatus
            ws.cell(row = index+1, column = 5).value = testDataResults[index].expectedError
            ws.cell(row = index+1, column = 6).value = testDataResults[index].actualOutput
            ws.cell(row = index+1, column = 7).value = testDataResults[index].actualHttpStatus
            ws.cell(row = index+1, column = 8).value = testDataResults[index].actualError
            ws.cell(row = index+1, column = 9).value = testDataResults[index].result
            ws.cell(row = index+1, column = 10).value = testDataResults[index].testSummary
            ws.cell(row = index+1, column = 11).value = testDataResults[index].testDescription
        wb.save(self.getResultsFilePath())


